
#openpyxl se usa para manejar una tabla excel
from cProfile import label
from cgitb import text
from tkinter import messagebox
from openpyxl import *

#tkinter se usa para crear una ventana grafica
from tkinter import *


class Formulario(object):
    def __init__(self) -> None:
        self.v = Tk()
        self.v.geometry("400x400+200+200")
        self.v.title("ventana prueba excel")
        self.opciones()
        self.v.mainloop()
        
    def opciones(self):
        # lista para almacenar los registros
        self.listaReg = []
        # se agragan Label, Entry, Button, Stringvar(simpre la inicial es mayuscula)
        self.lblCodigo = Label(self.v, text="codigo")
        self.lblCodigo.place(x=30,y=30)
        self.codigo = StringVar()
        self.eCodigo = Entry(self.v, textvariable=self.codigo)
        self.eCodigo.place(x=90,y=30)
            
        self.lblNombre = Label(self.v, text="nombre")
        self.lblNombre.place(x=30,y=100)
        self.nombre = StringVar()
        self.eNombre = Entry(self.v, textvariable=self.nombre)
        self.eNombre.place(x=90,y=100)
        
        self.btnIngresar = Button(text="ingrasar", command=lambda: self.ingresar())
        self.btnIngresar.place(x=30, y=150)
        
        self.btnSalir = Button(text="salir", command=lambda: self.salir ())
        self.btnSalir.place(x=100, y=150)
        
    def ingresar(self):
        reg = "{0};{1}\n".format(self.codigo.get(), self.nombre.get())
        self.listaReg.append(reg)
        self.codigo.set("")
        self.nombre.set("")
        messagebox.showinfo(title="ingresando", message="registro ingresado") 
    
    def salir(self):
        # grabo el archivo csv
        try:
            wb = load_workbook("listado.xlsx")
            ws = wb.active
        except:
            #primero se instancia un libro de excel y 
            # automaticamente una hoja de excel se instancia
            wb = Workbook()
            ws = wb.active
            ws["a1"] = "Codigo"
            ws["b1"] = "Nombre"
        
        for registro in self.listaReg:
            ws.append(registro.strip().split(";"))
        
        # grabo el excel
        wb.save("listado.xlsx")
        
        #arch = open(file="listado.csv", mode="a")
        #arch.writelines(self.listaReg)
        #arch.close()
        messagebox.showinfo(title="saliendo", message="terminado")
        exit() 
    
    
        
form = Formulario()